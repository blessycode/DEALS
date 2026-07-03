import os, re
from datetime import datetime
from openpyxl import load_workbook, Workbook

YEAR = "2026"
MONTHS = {1:"JAN",2:"FEB",3:"MAR",4:"APR",5:"MAY",6:"JUN",7:"JUL",8:"AUG",9:"SEP",10:"OCT",11:"NOV",12:"DEC"}

def clean(v):
    return "" if v is None else str(v).strip()

def row_text(row):
    return " ".join(clean(v) for v in row if clean(v)).upper()

def sheet_date(name):
    name = str(name).strip()
    if re.fullmatch(r"\d{8}", name) and name.endswith(YEAR):
        return datetime.strptime(name, "%d%m%Y")
    return None

def currency_from_title(text):
    t = text.upper()
    m = re.search(r"\(([^)]+)\)", t)
    if m:
        cur = m.group(1).strip().upper()
    elif "USD" in t:
        cur = "USD"
    elif "ZIG" in t:
        cur = "ZIG"
    elif "ZWG" in t:
        cur = "ZWG"
    elif "ZWL" in t:
        cur = "ZWL"
    else:
        cur = ""
    return "ZWL" if cur == "ZIG" else cur

def count_nonblank(row):
    return sum(1 for v in row if clean(v))

def is_section_row(row, rt):
    return count_nonblank(row) <= 6 and "TOTAL" not in rt and "TRANSACTION ID" not in rt and "BOOK" not in rt

def find_exact(row, names):
    names = {x.upper() for x in names}
    for i, v in enumerate(row):
        if clean(v).upper() in names:
            return i
    return None

def find_prefix(row, prefixes):
    prefixes = [p.upper() for p in prefixes]
    for i, v in enumerate(row):
        x = clean(v).upper()
        for p in prefixes:
            if x.startswith(p):
                return i
    return None

def headers_from(row, start):
    h = [clean(v) for v in row[start:]]
    while h and h[-1] == "":
        h.pop()
    return h if h else ["Transaction ID"]

def output_headers(headers, include_common_reference):
    extra_headers = ["CURRENCY"]
    if include_common_reference:
        extra_headers.append("Common Reference")
    extra_headers.append("DATE")
    return headers + extra_headers

def value_from_headers(data, headers, preferred_headers):
    normalized_preferred = {
        clean(header).upper().replace(" ", "")
        for header in preferred_headers
    }
    for i, header in enumerate(headers):
        normalized = clean(header).upper().replace(" ", "")
        if normalized in normalized_preferred and i < len(data):
            return data[i]
    return data[0] if data else ""

def common_reference(data, headers, dt, preferred_headers):
    reference_id = value_from_headers(data, headers, preferred_headers)
    return f"{reference_id}{dt.day}" if reference_id else ""

def get_sheet(out_wb, cache, name, headers, include_common_reference):
    if name not in cache:
        ws = out_wb.create_sheet(name)
        ws.append(output_headers(headers, include_common_reference))
        cache[name] = {
            "ws": ws,
            "headers": headers,
            "include_common_reference": include_common_reference,
        }
    return cache[name]["ws"], cache[name]["headers"]

def append_data(ws, row, start, headers, date_value, currency, dt, common_reference_headers=None):
    data = [clean(v) for v in row[start:start+len(headers)]]
    while len(data) < len(headers):
        data.append("")
    data.append(currency)
    if common_reference_headers:
        data.append(common_reference(data, headers, dt, common_reference_headers))
    data.append(date_value)
    ws.append(data)

def build_assets_workbook(file_path_or_stream):
    wb = load_workbook(file_path_or_stream, read_only=True, data_only=True)
    out = Workbook()
    out.remove(out.active)

    cache = {}
    mm_count = 0
    sec_count = 0

    for sheet_name in wb.sheetnames:
        dt = sheet_date(sheet_name)
        if not dt:
            continue

        ws = wb[sheet_name]
        month = MONTHS[dt.month]
        date_value = dt.strftime("%d/%m/%Y")

        section = None
        currency = ""

        mm_headers = None
        mm_start = None
        sec_headers = None
        sec_start = None

        for row in ws.iter_rows(values_only=True):
            row = list(row)
            rt = row_text(row)

            if not rt:
                continue

            if "TOTAL" in rt:
                continue

            if "LIABILITIES" in rt:
                section = None
                currency = ""
                continue

            if is_section_row(row, rt) and "HTM" in rt and "ASSET" in rt:
                section = "HTM"
                currency = currency_from_title(rt)
                continue

            if is_section_row(row, rt) and "HTM" not in rt and ("ASSET" in rt or "PLACEMENT" in rt or "MONEY MARKET" in rt):
                section = "MM"
                currency = currency_from_title(rt)
                continue

            if section == "MM":
                h = find_exact(row, {"TRANSACTION ID", "TRANSACTIONID"})
                if h is not None:
                    mm_start = h
                    mm_headers = headers_from(row, mm_start)
                    continue

                tx = find_prefix(row, ["MM"])
                if tx is not None:
                    if mm_headers is None:
                        mm_start = tx
                        mm_headers = [f"Column_{i+1}" for i in range(len(row[mm_start:]))]

                    out_ws, headers = get_sheet(out, cache, f"MM_{month}", mm_headers, True)
                    append_data(
                        out_ws,
                        row,
                        mm_start,
                        headers,
                        date_value,
                        currency,
                        dt,
                        {"TRANSACTION ID", "TRANSACTIONID"}
                    )
                    mm_count += 1

            elif section == "HTM":
                h = find_exact(row, {"BOOK"})
                if h is not None:
                    sec_start = h
                    sec_headers = headers_from(row, sec_start)
                    continue

                tx = find_prefix(row, ["9999"])
                if tx is not None:
                    if sec_headers is None:
                        sec_start = tx
                        sec_headers = [f"Column_{i+1}" for i in range(len(row[sec_start:]))]

                    out_ws, headers = get_sheet(out, cache, f"SECURITIES_{month}", sec_headers, True)
                    append_data(
                        out_ws,
                        row,
                        sec_start,
                        headers,
                        date_value,
                        currency,
                        dt,
                        {"ID"}
                    )
                    sec_count += 1

    return out, mm_count, sec_count


def main():
    from tkinter import Tk, filedialog

    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title="Select workbook",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    if not file_path:
        print("No file selected.")
        return

    out, mm_count, sec_count = build_assets_workbook(file_path)
    output = os.path.join(os.path.dirname(file_path), "Combined_Assets_By_Month_2026.xlsx")
    out.save(output)

    print("DONE")
    print("MM rows:", mm_count)
    print("SECURITIES rows:", sec_count)
    print(output)

if __name__ == "__main__":
    main()
